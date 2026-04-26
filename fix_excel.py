"""
Script na opravu Excel sloupců - přejmenuj 'Popis' na 'tagy' a přidej 'barva_satu'
Spusť v stejné složce jako Data.xlsx
"""
from openpyxl import load_workbook
import sys

try:
    # Najdi soubor Data.xlsx
    wb = load_workbook('Data.xlsx')
    ws = wb.active
    
    # Přečti header
    header = [cell.value for cell in ws[1]]
    print(f"Aktuální header: {header}")
    
    # Najdi a přejmenuj sloupce
    mapping = {
        'Popis': 'tagy',
        'popis': 'tagy',
        'POPIS': 'tagy'
    }
    
    for col_idx, cell_value in enumerate(header, 1):
        if cell_value in mapping:
            ws.cell(1, col_idx).value = mapping[cell_value]
            print(f"✓ Přejmenován: '{cell_value}' -> '{mapping[cell_value]}'")
    
    # Kontrola, jestli existuje 'barva_satu'
    header = [cell.value for cell in ws[1]]
    if 'barva_satu' not in header and 'barva_satu' not in [str(x).lower() for x in header]:
        # Přidej do posledního sloupce + 1
        last_col = len(header) + 1
        ws.cell(1, last_col).value = 'barva_satu'
        print(f"✓ Přidán sloupec: 'barva_satu'")
    
    # Ulož
    wb.save('Data.xlsx')
    print("\n✅ Excel opraveno a uloženo!")
    print("Nyní spusť znovu: python pinterest_automation.py")
    
except FileNotFoundError:
    print("❌ Soubor Data.xlsx nenalezen v aktuální složce")
    print("Ujisti se, že jsi v správné složce a soubor se jmenuje 'Data.xlsx'")
    sys.exit(1)
except Exception as e:
    print(f"❌ Chyba: {e}")
    sys.exit(1)
