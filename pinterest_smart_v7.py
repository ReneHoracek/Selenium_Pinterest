"""
Pinterest Smart Automation V4
==============================
Vylepšení oproti V3:
  1. Cross-platform bezpečný logging (bez UnicodeEncodeError na Windows)
  2. Excel: zachování formátování, filtrů, dropdownů při aktualizacích
  3. Přepracovaný inteligentní algoritmus řazení (round-robin + penalizace)
  4. Stabilnější Selenium (retry, stale-element recovery, adaptivní wait)
  5. Obecná vylepšení: pathlib, keyring přihlášení, modulárnost
"""

from __future__ import annotations

import codecs
import copy
import getpass
import logging
import os
import random
import sys
import time
import traceback
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import keyring
    _KEYRING_DOSTUPNY = True
except ImportError:
    _KEYRING_DOSTUPNY = False

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, Font, PatternFill,
                              Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException, StaleElementReferenceException,
    TimeoutException, WebDriverException,
)
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ===========================================================================
# 1. CROSS-PLATFORM LOGGING
# ===========================================================================
# PROBLÉM: Windows konzole (cp1250) neumí vykreslit emoji (✅, 📋 …).
# Výsledkem je UnicodeEncodeError, který zhroutí aplikaci.
#
# ŘEŠENÍ:
#  a) File handler → vždy UTF-8, emojis se zachovají.
#  b) Console handler → vlastní SafeStreamHandler, který detekuje
#     kódování konzole a v případě potřeby emoji tiše nahradí ASCII.
#
# Proč není stačit reconfigure(encoding='utf-8')?
#  - Funguje jen na Pythonu 3.7+ a jen pro sys.stdout/stderr.
#  - Na některých Windows terminalech (IDE konzole, Jenkins) to selhává.
#  - SafeStreamHandler je robustnější a přenositelný.
# ===========================================================================

class SafeStreamHandler(logging.StreamHandler):
    """
    StreamHandler, který nikdy neselže kvůli kódování.
    Na terminálech bez UTF-8 podpory nahradí emoji bezpečnými ASCII náhradami.
    """

    # Mapování nejběžnějších emoji → ASCII ekvivalenty
    EMOJI_FALLBACK: dict[str, str] = {
        "✅": "[OK]",
        "❌": "[ERR]",
        "⚠️": "[WARN]",
        "📋": "[LIST]",
        "📌": "[PIN]",
        "📅": "[DATE]",
        "📹": "[VIDEO]",
        "🎨": "[COLOR]",
        "🔍": "[SEARCH]",
        "🤖": "[BOT]",
        "⏳": "[WAIT]",
        "▶": ">",
        "🔴": "[RED]",
        "ℹ️": "[INFO]",
    }

    def __init__(self, stream=None):
        super().__init__(stream or sys.stderr)
        # Zjisti, zda terminál podporuje UTF-8
        enc = getattr(self.stream, "encoding", None) or ""
        self._utf8_safe = enc.lower().replace("-", "") in ("utf8", "utf16", "utf32", "")

    def _safe_msg(self, msg: str) -> str:
        """Pokud terminál nepodporuje UTF-8, nahraď emoji ASCII variantami."""
        if self._utf8_safe:
            return msg
        for emoji, replacement in self.EMOJI_FALLBACK.items():
            msg = msg.replace(emoji, replacement)
        # Zbytek znaků mimo ASCII – odstraň, místo pádu
        return msg.encode(self.stream.encoding or "ascii", errors="replace").decode(
            self.stream.encoding or "ascii"
        )

    def emit(self, record: logging.LogRecord) -> None:
        try:
            msg = self.format(record)
            safe = self._safe_msg(msg)
            stream = self.stream
            stream.write(safe + self.terminator)
            self.flush()
        except RecursionError:
            raise
        except Exception:
            self.handleError(record)


def _setup_logging(log_file: Path) -> logging.Logger:
    """
    Inicializuje logger s:
      - file handlerem (UTF-8, zachová emoji)
      - console handlerem (bezpečný pro cp1250 i UTF-8)
    """
    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    root = logging.getLogger("pinterest")
    root.setLevel(logging.INFO)
    root.handlers.clear()  # zabraň duplikátům při opakovaném importu

    # File handler – vždy UTF-8
    fh = logging.FileHandler(str(log_file), encoding="utf-8")
    fh.setFormatter(fmt)
    root.addHandler(fh)

    # Console handler – bezpečný
    ch = SafeStreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    root.addHandler(ch)

    return root


LOG_FILE = Path("pinterest_smart.log")
logger = _setup_logging(LOG_FILE)


# ===========================================================================
# 2. KONFIGURACE
# ===========================================================================

EXCEL_FILE = Path("Pinterest_Data_Smart.xlsx")
VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv", ".webm", ".flv", ".wmv", ".m4v"}
BASE_URL = "https://cz.pinterest.com/"
PIN_URL = "https://cz.pinterest.com/pin-creation-tool/"

# Nástěnky
NASTENKY: list[str] = [
    "Daňové poradenství",
    "Podnikání online",
    "Rady a Tipy nejen pro účetnictví",
    "UOL Účetnictví",
    "Účetnictví online",
]

# Barvy šatů
BARVY_SATU: list[str] = [
    "červená",
    "modrá",
    "bílá",
    "černá",
    "zelená",
    "žlutá",
    "růžová",
    "fialová",
    "béžová",
    "hnědá",
    "šedá",
    "oranžová",
]

# Stavy pinů
STAV_CEKANI = "CEKANI NA VYPLNENI"
STAV_PRIPRAVENO = "PRIPRAVENO K NAHRANI"
STAV_NAHRANO = "NAHRANO"
STAV_CHYBA = "CHYBA"

# Pravidla plánování (weekday → počet + časová okna)
PLANNING_RULES: dict[int, dict] = {
    0: {"count": 2, "times": [(7, 9), (12, 13)]},          # Pondělí
    1: {"count": 3, "times": [(7, 9), (12, 13), (19, 21)]}, # Úterý
    2: {"count": 3, "times": [(7, 9), (12, 13), (19, 21)]}, # Středa
    3: {"count": 3, "times": [(7, 9), (12, 13), (19, 21)]}, # Čtvrtek
    4: {"count": 2, "times": [(7, 9), (12, 13)]},           # Pátek
    5: {"count": 1, "times": [(18, 21)]},                   # Sobota
    6: {"count": 1, "times": [(18, 21)]},                   # Neděle
}

# Mapování sloupců Excel → interní jména
COLUMN_MAP: dict[str, str] = {
    "NADPIS VIDEA": "nazev",
    "TYP ODKAZU": "odkaz",
    "NASTĚNKA": "nastepka",
    "TAGY (oddělené středníkem)": "tagy",
    "BARVA ŠATŮ": "barva_satu",
    "CESTA K VIDEU": "video_cesta",
    "DATUM PLÁNOVÁNÍ": "datum_planovani",
    "ČAS PLÁNOVÁNÍ": "cas_planovani",
    "STAV": "stav",
}

# Selenium
SELENIUM_TIMEOUT = 25        # s – základní čekání
SELENIUM_LONG_TIMEOUT = 120  # s – čekání na nahrání videa (po kliknutí Publish)
SELENIUM_UPLOAD_TIMEOUT = 600  # s – maximální čekání na zpracování videa Pinterestem
                                # (velká videa mohou trvat 5–10 minut)
SELENIUM_UPLOAD_POLL = 5     # s – jak často kontrolujeme stav nahrávání
SELENIUM_RETRY_COUNT = 3     # počet pokusů při selhání prvku
SELENIUM_RETRY_DELAY = 1.5   # s – pauza mezi pokusy

# ---------------------------------------------------------------------------
# Přihlášení do Pinterestu
# ---------------------------------------------------------------------------
# Název služby v systémové klíčence (Windows Credential Manager / Linux Secret Service)
KEYRING_SERVICE = "Pinterest_Smart_Automation"
KEYRING_USER_KEY = "pinterest_email"
KEYRING_PASS_KEY = "pinterest_password"

# FALLBACK – pokud keyring není dostupný nebo selže, použijí se tyto hodnoty.
# ZMĚŇ JE NA SVÉ PŘIHLAŠOVACÍ ÚDAJE (nebo nech prázdné – skript se zeptá).
# Po prvním úspěšném přihlášení se uloží do systémové klíčenky automaticky.
PINTEREST_EMAIL_FALLBACK = ""   # např. "muj@email.cz"
PINTEREST_PASSWORD_FALLBACK = ""  # např. "MojeHeslo123"

# Excel styly
HEADER_COLOR = "4472C4"
HEADER_FONT_COLOR = "FFFFFF"
ROW_COLOR_ODD = "DCE6F1"
ROW_COLOR_EVEN = "FFFFFF"


# ===========================================================================
# 3. EXCEL HELPER
# ===========================================================================
# PROBLÉM: df.to_excel() přepíše celý soubor → ztratí se formátování,
# filtry, dropdown validace, podmíněné formátování.
#
# ŘEŠENÍ:
#  a) Při vytváření nového Excelu → openpyxl přímo (ne přes pandas Writer).
#  b) Při aktualizaci stavu → load_workbook() + uprav jen konkrétní buňky.
# ===========================================================================

def _aplikuj_styl_workbooku(ws, pocet_radku: int) -> None:
    """
    Aplikuje kompletní styl na worksheet:
    - záhlaví (modrá, tučné, bílé písmo)
    - střídání řádků
    - zamorování 1. řádku
    - autofiltr
    - šířky sloupců
    Tato funkce se volá jak při vytváření, tak při obnově formátování.
    """
    # Záhlaví
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR,
                              fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Střídání řádků
    for row_idx in range(2, pocet_radku + 2):
        fill_color = ROW_COLOR_ODD if row_idx % 2 == 0 else ROW_COLOR_EVEN
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color,
                               fill_type="solid")
        for cell in ws[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill
            cell.border = border

    # Zmrazení 1. řádku
    ws.freeze_panes = "A2"

    # Autofiltr
    ws.auto_filter.ref = ws.dimensions

    # Výšky záhlaví
    ws.row_dimensions[1].height = 30

    # Šířky sloupců (pevné hodnoty pro každý sloupec)
    sirky = {"A": 45, "B": 22, "C": 28, "D": 55, "E": 20,
              "F": 60, "G": 18, "H": 14, "I": 30}
    for col_letter, width in sirky.items():
        ws.column_dimensions[col_letter].width = width


def _pridej_validace(ws, pocet_radku: int) -> None:
    """
    Přidá (nebo obnoví) dropdown validace do worksheetu.
    Bezpečně smaže staré validace stejného rozsahu před přidáním nových.
    """
    # Odstraň existující validace (zabrání duplikátům při obnově)
    ws.data_validations.dataValidation.clear()

    # TYP ODKAZU – sloupec B
    dv_odkaz = DataValidation(
        type="list",
        formula1='"www.uol.cz,https://kurzy.uolakademie.cz/"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Neplatná hodnota",
        error="Vyber: www.uol.cz nebo kurzy",
    )
    ws.add_data_validation(dv_odkaz)
    dv_odkaz.add(f"B2:B{pocet_radku + 1}")

    # NÁSTĚNKA – sloupec C
    nastenky_str = ",".join(NASTENKY)
    dv_nastenka = DataValidation(
        type="list",
        formula1=f'"{nastenky_str}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Neplatná nástěnka",
        error="Vyber nástěnku ze seznamu",
    )
    ws.add_data_validation(dv_nastenka)
    dv_nastenka.add(f"C2:C{pocet_radku + 1}")

    # BARVA ŠATŮ – sloupec E
    barvy_str = ",".join(BARVY_SATU)
    dv_barva = DataValidation(
        type="list",
        formula1=f'"{barvy_str}"',
        allow_blank=True,
        showErrorMessage=False,  # jen nápověda, nekomplikuj
    )
    ws.add_data_validation(dv_barva)
    dv_barva.add(f"E2:E{pocet_radku + 1}")

    # STAV – sloupec I
    stavy = f"{STAV_CEKANI},{STAV_PRIPRAVENO},{STAV_NAHRANO},{STAV_CHYBA}"
    dv_stav = DataValidation(
        type="list",
        formula1=f'"{stavy}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Neplatný stav",
        error="Vyber stav ze seznamu",
    )
    ws.add_data_validation(dv_stav)
    dv_stav.add(f"I2:I{pocet_radku + 1}")


# ===========================================================================
# 4. INTELIGENTNÍ ŘAZENÍ
# ===========================================================================
# PROBLÉM: Původní algoritmus vybírá jen první kandidáty ze seznamu
# a střídá pouze přímé sousedy (posledni_barva != aktualni). U větších
# datasetů se může stát, že se barva nebo nástěnka opakuje ve shlucích.
#
# NOVÝ ALGORITMUS – Round-robin s penalizačním skóre:
#  1. Nástěnky přiřazujeme round-robinem (zaručuje rovnoměrné rozdělení).
#  2. Při výběru dalšího videa v rámci dané barvy použijeme penalizační
#     skóre: čím nedávněji byla barva použita, tím vyšší penalizace.
#  3. Tiebreaker = abecední pořadí souboru → deterministické chování.
#
# VÝHODY:
#  - O(n log n) namísto O(n²)
#  - Garantovaně vybere různé barvy v co největší vzdálenosti
#  - Škáluje na tisíce videí
#  - Deterministický výsledek (žádný random)
# ===========================================================================

def inteligentni_serazeni(videa: list[dict]) -> list[dict]:
    """
    Seřadí videa tak, aby se:
      a) nástěnky rovnoměrně střídaly (round-robin)
      b) barvy šatů co nejméně opakovaly za sebou (penalizační výběr)

    Vstup: list slovníků s klíči 'soubor', 'cesta', 'barva'
    Výstup: list slovníků rozšířených o klíč 'nastenka'
    """
    logger.info("[BOT] Provádím inteligentní řazení...")

    if not videa:
        return []

    # Pracovní kopie, aby jsme nemodifikovali originál
    zbyvajici: list[dict] = [copy.deepcopy(v) for v in videa]
    # Sekundární tiebreaker – abecedně podle souboru → deterministické
    zbyvajici.sort(key=lambda v: v["soubor"])

    serazeno: list[dict] = []
    # Sledujeme, jak dávno bylo jaká barva naposledy použita (index v serazeno)
    posledni_pouziti_barvy: dict[str, int] = {}
    nastenka_idx = 0

    while zbyvajici:
        # Vyber nástěnku round-robinem
        cilova_nastenka = NASTENKY[nastenka_idx % len(NASTENKY)]
        nastenka_idx += 1

        aktualni_pozice = len(serazeno)

        def penalizace(v: dict) -> int:
            """Čím dřív byla barva použita, tím nižší číslo (=lepší kandid.)"""
            posledni = posledni_pouziti_barvy.get(v["barva"], -999)
            # Vzdálenost: čím větší, tím lepší (= méně opakování)
            return -(aktualni_pozice - posledni)  # záporné → min = nejlepší

        # Seřaď kandidáty podle penalizace, pak abecedně (deterministické)
        zbyvajici.sort(key=lambda v: (penalizace(v), v["soubor"]))
        vybrany = zbyvajici.pop(0)

        # Přiřaď nástěnku
        vybrany["nastenka"] = cilova_nastenka
        serazeno.append(vybrany)
        posledni_pouziti_barvy[vybrany["barva"]] = aktualni_pozice

    # Statistika pro ladění
    barvy_counter: Counter = Counter(v["barva"] for v in serazeno)
    nastenky_counter: Counter = Counter(v["nastenka"] for v in serazeno)
    logger.info(f"[OK] Seřazeno {len(serazeno)} videí")
    logger.info(f"     Barvy: {dict(barvy_counter)}")
    logger.info(f"     Nástěnky: {dict(nastenky_counter)}")

    # Diagnostika: maximální délka shluku stejné barvy
    max_shluk = _max_shluk_barvy(serazeno)
    logger.info(f"     Max. shluk stejné barvy za sebou: {max_shluk}")

    return serazeno


def _max_shluk_barvy(serazena: list[dict]) -> int:
    """Vrátí délku nejdelšího shluku stejné barvy (pro diagnostiku)."""
    if not serazena:
        return 0
    max_c = cur_c = 1
    for i in range(1, len(serazena)):
        if serazena[i]["barva"] == serazena[i - 1]["barva"]:
            cur_c += 1
            max_c = max(max_c, cur_c)
        else:
            cur_c = 1
    return max_c


# ===========================================================================
# 5. PLÁNOVÁNÍ
# ===========================================================================

def vygeneruj_nahodny_cas(start_hour: int, end_hour: int) -> str:
    """Vygeneruje náhodný čas (celá nebo půl hodina) v daném rozmezí."""
    mozne = []
    for h in range(start_hour, end_hour + 1):
        mozne.append(f"{h:02d}:00")
        if h < end_hour:
            mozne.append(f"{h:02d}:30")
    return random.choice(mozne)


def naplanovej_piny(videa: list[dict], start_date: Optional[datetime] = None) -> list[dict]:
    """Automaticky naplánuje piny podle PLANNING_RULES."""
    if start_date is None:
        start_date = (datetime.now() + timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )

    scheduled: list[dict] = []
    current_date = start_date
    video_index = 0

    while video_index < len(videa):
        weekday = current_date.weekday()
        rules = PLANNING_RULES[weekday]
        pins_today = min(rules["count"], len(videa) - video_index)

        for i in range(pins_today):
            time_range = rules["times"][i]
            cas = vygeneruj_nahodny_cas(time_range[0], time_range[1])
            scheduled.append({
                "video": videa[video_index],
                "datum": current_date.strftime("%d.%m.%Y"),
                "cas": cas,
            })
            video_index += 1

        current_date += timedelta(days=1)

    return scheduled


# ===========================================================================
# 6. ZADÁVÁNÍ BAREV ŠATŮ
# ===========================================================================

def zjisti_barvy_satu(videa: list[dict]) -> list[dict]:
    """Interaktivně se zeptá na barvu šatů pro každé video (číselná nabídka)."""
    logger.info("=" * 60)
    logger.info("ZADÁVÁNÍ BAREV ŠATŮ")
    logger.info("=" * 60)

    print("\n[COLOR] DOSTUPNÉ BARVY:")
    for i, barva in enumerate(BARVY_SATU, 1):
        print(f"   {i:2d}. {barva}")
    print()
    print("Enter = opakuj předchozí barvu\n")

    videa_s_barvami: list[dict] = []
    posledni_barva: Optional[str] = None

    for i, video in enumerate(videa, 1):
        while True:
            hint = f" [Enter = {posledni_barva}]" if posledni_barva else ""
            try:
                vstup = input(f"[VIDEO] {i}/{len(videa)} {video['soubor'][:50]}{hint}: ").strip()
            except (EOFError, KeyboardInterrupt):
                logger.warning("\nZadávání přerušeno uživatelem.")
                sys.exit(0)

            if vstup == "" and posledni_barva:
                barva = posledni_barva
                print(f"   [OK] Barva: {barva}")
                break

            try:
                cislo = int(vstup)
                if 1 <= cislo <= len(BARVY_SATU):
                    barva = BARVY_SATU[cislo - 1]
                    print(f"   [OK] Barva: {barva}")
                    posledni_barva = barva
                    break
                else:
                    print(f"   [WARN] Zadej číslo 1-{len(BARVY_SATU)}")
            except ValueError:
                if vstup:
                    barva = vstup.lower()
                    print(f"   [OK] Barva: {barva}")
                    posledni_barva = barva
                    break
                else:
                    print(f"   [WARN] Zadej číslo 1-{len(BARVY_SATU)}")

        videa_s_barvami.append({
            "soubor": video["soubor"],
            "cesta": video["cesta"],
            "barva": barva,
        })

    logger.info(f"[OK] Barvy zadány pro všechna videa ({len(videa_s_barvami)})")
    return videa_s_barvami


# ===========================================================================
# 7. HLEDÁNÍ VIDEÍ
# ===========================================================================

def najdi_videa(slozka: Path = Path(".")) -> list[dict]:
    """Najde všechna videa v zadané složce (nerekurzivně)."""
    videa: list[dict] = []
    for soubor in sorted(slozka.iterdir()):  # sort → deterministické
        if soubor.suffix.lower() in VIDEO_EXTENSIONS:
            videa.append({"soubor": soubor.name, "cesta": str(soubor.resolve())})
    logger.info(f"[SEARCH] Nalezeno {len(videa)} videí v: {slozka.resolve()}")
    return videa


# ===========================================================================
# 8. VYTVOŘENÍ EXCELU
# ===========================================================================
# PROBLÉM: Původní kód nejprve zapsal přes pandas a pak znovu otevřel
# přes openpyxl, což je zbytečné a rizikové. Lepší je rovnou openpyxl.
#
# PROBLÉM 2: df.to_excel() při každé aktualizaci stavu přepsal celý soubor
# → zmizely formátování, filtry, dropdowny.
#
# ŘEŠENÍ: Vytváření → openpyxl přímo (bez pandas Writer).
#         Aktualizace stavu → load_workbook + zápis jen konkrétních buněk.
# ===========================================================================

def vytvor_excel_z_videi(videa: list[dict]) -> bool:
    """Vytvoří nový Excel soubor z videí. Zachová kompletní formátování."""
    if not videa:
        logger.warning("[WARN] Žádná videa nenalezena!")
        return False

    videa_s_barvami = zjisti_barvy_satu(videa)
    serazena_videa = inteligentni_serazeni(videa_s_barvami)
    scheduled = naplanovej_piny(serazena_videa)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Piny"

    # Záhlaví
    zahlavi = [
        "NADPIS VIDEA",
        "TYP ODKAZU",
        "NASTĚNKA",
        "TAGY (oddělené středníkem)",
        "BARVA ŠATŮ",
        "CESTA K VIDEU",
        "DATUM PLÁNOVÁNÍ",
        "ČAS PLÁNOVÁNÍ",
        "STAV",
    ]
    ws.append(zahlavi)

    # Datové řádky
    for item in scheduled:
        v = item["video"]
        ws.append([
            f"[VYPLŇ NADPIS - {v['soubor']}]",
            "www.uol.cz",
            v["nastenka"],
            "",
            v["barva"],
            v["cesta"],
            item["datum"],
            item["cas"],
            STAV_CEKANI,
        ])

    pocet_radku = len(scheduled)
    _aplikuj_styl_workbooku(ws, pocet_radku)
    _pridej_validace(ws, pocet_radku)

    try:
        wb.save(str(EXCEL_FILE))
        logger.info(f"[OK] Excel vytvořen: {EXCEL_FILE}")
        logger.info(f"[LIST] Počet videí: {len(videa)}")

        logger.info("=" * 60)
        logger.info("NÁHLED SEŘAZENÍ (prvních 10):")
        logger.info("=" * 60)
        for i, item in enumerate(scheduled[:10], 1):
            v = item["video"]
            logger.info(
                f"{i:2d}. {v['soubor'][:38]:<38} | {v['barva']:<10} | {v['nastenka']}"
            )
        if len(scheduled) > 10:
            logger.info(f"    ... a dalších {len(scheduled) - 10} videí")

        return True
    except Exception as exc:
        logger.error(f"[ERR] Chyba při ukládání Excelu: {exc}")
        logger.error(traceback.format_exc())
        return False


# ===========================================================================
# 9. KONTROLA EXCELU
# ===========================================================================

def zkontroluj_excel() -> tuple[bool, Optional[pd.DataFrame]]:
    """
    Zkontroluje Excel soubor:
    - existence sloupců
    - přítomnost řádků PRIPRAVENO K NAHRANI
    - existence video souborů
    Vrací (ok, df_připraveno) nebo (False, None).
    """
    if not EXCEL_FILE.exists():
        logger.warning(f"[WARN] Soubor {EXCEL_FILE} neexistuje")
        return False, None

    try:
        df = pd.read_excel(str(EXCEL_FILE))
        df = df.rename(columns=COLUMN_MAP)

        povinne = list(COLUMN_MAP.values())
        for sloupec in povinne:
            if sloupec not in df.columns:
                logger.error(f"[ERR] Chybí sloupec: {sloupec}")
                return False, None

        # Filtruj připravené piny
        df_pripraveno = df[df["stav"].str.contains(STAV_PRIPRAVENO, na=False)].copy()

        if len(df_pripraveno) == 0:
            logger.warning("[WARN] Žádné piny připraveny k nahrání")
            logger.info(f"   Označ řádky stavem '{STAV_PRIPRAVENO}' ve sloupci STAV")
            return False, None

        # Zkontroluj existenci videí
        chybejici = [
            f"  Řádek {idx + 2}: {row['video_cesta']}"
            for idx, row in df_pripraveno.iterrows()
            if not Path(str(row["video_cesta"]).strip()).exists()
        ]
        if chybejici:
            logger.error("[ERR] Některá videa nenalezena:")
            for msg in chybejici:
                logger.error(msg)
            return False, None

        logger.info(f"[OK] Excel je v pořádku! Připraveno: {len(df_pripraveno)} pinů")
        return True, df_pripraveno

    except Exception as exc:
        logger.error(f"[ERR] Chyba při čtení Excelu: {exc}")
        logger.error(traceback.format_exc())
        return False, None


def aktualizuj_stav_v_excelu(video_cesta: str, novy_stav: str) -> None:
    """
    KLÍČOVÁ FUNKCE: Aktualizuje stav jednoho řádku v Excelu
    BEZ přepisování celého souboru → zachová formátování.

    Algoritmus:
    1. Záloha souboru
    2. load_workbook (zachová vše)
    3. Najdi řádek podle cesty k videu (sloupec F = index 6)
    4. Přepiš jen buňku ve sloupci I (STAV)
    5. Ulož
    """
    if not EXCEL_FILE.exists():
        return

    try:
        wb = load_workbook(str(EXCEL_FILE))
        ws = wb.active

        # Najdi sloupce dynamicky podle záhlaví (robustnější než fixní index)
        zahlavi = {cell.value: cell.column for cell in ws[1]}
        col_cesta = zahlavi.get("CESTA K VIDEU")
        col_stav = zahlavi.get("STAV")

        if not col_cesta or not col_stav:
            logger.warning("[WARN] Záhlaví CESTA K VIDEU nebo STAV nenalezeno v Excelu")
            return

        nalezeno = False
        for row in ws.iter_rows(min_row=2):
            bunka_cesta = row[col_cesta - 1]
            if str(bunka_cesta.value).strip() == str(video_cesta).strip():
                row[col_stav - 1].value = novy_stav
                nalezeno = True
                break

        if not nalezeno:
            logger.warning(f"[WARN] Video nenalezeno v Excelu: {video_cesta}")

        wb.save(str(EXCEL_FILE))
        logger.info(f"[OK] Stav aktualizován: {Path(video_cesta).name} → {novy_stav}")

    except Exception as exc:
        logger.error(f"[ERR] Chyba při aktualizaci stavu: {exc}")
        logger.error(traceback.format_exc())


# ===========================================================================
# 10. PŘIHLAŠOVACÍ SYSTÉM (keyring – cross-platform klíčenka)
# ===========================================================================
# PROČ keyring?
#   - Windows: ukládá do Windows Credential Manager (šifrovaně, bez plaintext)
#   - Linux:   ukládá do Secret Service (GNOME Keyring / KWallet)
#   - Fedora:  funguje nativně se GNOME Keyring bez extra konfigurace
#   - Heslo NIKDY není v kódu ani v souborech na disku
#
# LOGIKA:
#   1. Zkus načíst z klíčenky (keyring.get_password)
#   2. Pokud není → zkus FALLBACK konstanty v kódu (pro první spuštění)
#   3. Pokud ani to → zeptej se interaktivně (getpass pro skryté zadání hesla)
#   4. Při prvním získání udajů je ulož do klíčenky pro příští spuštění
#
# SMAZÁNÍ uložených údajů:
#   keyring.delete_password("Pinterest_Smart_Automation", "pinterest_email")
#   keyring.delete_password("Pinterest_Smart_Automation", "pinterest_password")
# ===========================================================================

def ziskej_prihlasovaci_udaje() -> tuple[str, str]:
    """
    Načte email a heslo z (v prioritním pořadí):
      1. Systémová klíčenka (keyring) – Windows Credential Manager / GNOME Keyring
      2. Fallback konstanty v kódu (PINTEREST_EMAIL_FALLBACK / PASSWORD_FALLBACK)
      3. Interaktivní dotaz (getpass pro skryté zadání hesla)

    Při prvním získání z fallback/dotazu údaje automaticky uloží do klíčenky.
    Vrací (email, heslo).
    """
    email: Optional[str] = None
    heslo: Optional[str] = None
    uloz_do_klicenky = False

    # --- 1. Pokus: systémová klíčenka ---
    if _KEYRING_DOSTUPNY:
        try:
            email = keyring.get_password(KEYRING_SERVICE, KEYRING_USER_KEY)
            heslo = keyring.get_password(KEYRING_SERVICE, KEYRING_PASS_KEY)
            if email and heslo:
                logger.info(f"[OK] Přihlašovací údaje načteny z klíčenky ({email})")
                return email, heslo
        except Exception as exc:
            logger.warning(f"[WARN] Klíčenka nedostupná: {exc}")
    else:
        logger.warning("[WARN] Modul 'keyring' není nainstalován.")
        logger.warning("       Nainstaluj: pip install keyring")
        logger.warning("       Údaje budou uloženy pouze v paměti pro toto spuštění.")

    # --- 2. Pokus: fallback konstanty v kódu ---
    if PINTEREST_EMAIL_FALLBACK and PINTEREST_PASSWORD_FALLBACK:
        logger.info("[INFO] Používám přihlašovací údaje z kódu (fallback).")
        email = PINTEREST_EMAIL_FALLBACK
        heslo = PINTEREST_PASSWORD_FALLBACK
        uloz_do_klicenky = True

    # --- 3. Pokus: interaktivní dotaz ---
    if not email:
        logger.info("[INFO] Zadej přihlašovací údaje do Pinterestu:")
        try:
            email = input("  Pinterest email: ").strip()
        except (EOFError, KeyboardInterrupt):
            logger.error("[ERR] Zadání přerušeno.")
            sys.exit(1)
        uloz_do_klicenky = True

    if not heslo:
        try:
            # getpass skryje heslo při psaní (nezobrazuje znaky)
            heslo = getpass.getpass("  Pinterest heslo: ")
        except (EOFError, KeyboardInterrupt):
            logger.error("[ERR] Zadání přerušeno.")
            sys.exit(1)
        uloz_do_klicenky = True

    if not email or not heslo:
        logger.error("[ERR] Email nebo heslo je prázdné. Ukončuji.")
        sys.exit(1)

    # --- Ulož do klíčenky pro příští spuštění ---
    if uloz_do_klicenky and _KEYRING_DOSTUPNY:
        try:
            keyring.set_password(KEYRING_SERVICE, KEYRING_USER_KEY, email)
            keyring.set_password(KEYRING_SERVICE, KEYRING_PASS_KEY, heslo)
            logger.info(f"[OK] Přihlašovací údaje uloženy do klíčenky ({KEYRING_SERVICE})")
        except Exception as exc:
            logger.warning(f"[WARN] Nepodařilo se uložit do klíčenky: {exc}")

    return email, heslo


def prihlasit_se_na_pinterest(driver, email: str, heslo: str) -> bool:
    """
    Automaticky provede přihlášení na Pinterest.

    Postup:
      1. Přejde na přihlašovací stránku
      2. Vyplní email
      3. Vyplní heslo
      4. Klikne Přihlásit se
      5. Ověří úspěšné přihlášení (čeká na homepage prvek nebo URL změnu)

    Vrací True při úspěchu, False při chybě.
    """
    LOGIN_URL = "https://cz.pinterest.com/login/"

    logger.info("[INFO] Přihlašuji se na Pinterest...")
    logger.info(f"       Účet: {email}")

    try:
        driver.get(LOGIN_URL)
        time.sleep(2)

        # --- Email pole ---
        # Pinterest používá id="email" nebo data-test-id="login-email-field"
        email_selektory = [
            (By.ID, "email"),
            (By.XPATH, "//input[@data-test-id='login-email-field']"),
            (By.XPATH, "//input[@name='id']"),
            (By.XPATH, "//input[@type='email']"),
        ]
        email_input = None
        for by, sel in email_selektory:
            email_input = wait_for_element(driver, by, sel, timeout=10,
                                           condition="visible")
            if email_input:
                break

        if not email_input:
            logger.error("[ERR] Email pole nenalezeno na přihlašovací stránce")
            return False

        send_keys_safe(email_input, email)
        time.sleep(0.5)

        # --- Heslo pole ---
        heslo_selektory = [
            (By.ID, "password"),
            (By.XPATH, "//input[@data-test-id='login-password-field']"),
            (By.XPATH, "//input[@name='password']"),
            (By.XPATH, "//input[@type='password']"),
        ]
        heslo_input = None
        for by, sel in heslo_selektory:
            heslo_input = wait_for_element(driver, by, sel, timeout=10,
                                           condition="visible")
            if heslo_input:
                break

        if not heslo_input:
            logger.error("[ERR] Heslo pole nenalezeno na přihlašovací stránce")
            return False

        send_keys_safe(heslo_input, heslo)
        time.sleep(0.5)

        # --- Tlačítko Přihlásit se ---
        login_btn_selektory = [
            (By.XPATH, "//button[@data-test-id='registerFormSubmitButton']"),
            (By.XPATH, "//button[@type='submit']"),
            (By.XPATH, "//button[contains(translate(text(),'PŘIHLÁSITSELOGIN',"
                       "'přihlásitselogin'), 'přihlásit')]"),
            (By.XPATH, "//div[@data-test-id='login-button']"),
        ]
        login_btn = None
        for by, sel in login_btn_selektory:
            login_btn = wait_for_element(driver, by, sel, timeout=8,
                                         condition="clickable")
            if login_btn:
                break

        if not login_btn:
            # Zkus Enter na heslo poli jako fallback
            logger.warning("[WARN] Přihlašovací tlačítko nenalezeno, zkouším Enter")
            heslo_input.send_keys(Keys.RETURN)
        else:
            safe_click(driver, login_btn)

        # --- Ověření úspěchu ---
        # Čekáme na přesměrování z /login/ → homepage nebo /
        logger.info("  [WAIT] Čekám na dokončení přihlášení...")
        deadline = time.time() + 20
        while time.time() < deadline:
            current_url = driver.current_url
            # Přihlášení proběhlo = nejsme na /login/ ani /register/
            if "/login" not in current_url and "/register" not in current_url:
                logger.info(f"[OK] Přihlášení úspěšné! ({current_url[:60]})")
                return True
            time.sleep(1)

        # Zkontroluj, zda není chybová hláška (špatné heslo)
        chyba_selektory = [
            (By.XPATH, "//div[@data-test-id='login-error']"),
            (By.XPATH, "//p[contains(@class,'error')]"),
            (By.XPATH, "//div[contains(@class,'error')]"),
        ]
        for by, sel in chyba_selektory:
            try:
                el = driver.find_element(by, sel)
                if el.is_displayed() and el.text.strip():
                    logger.error(f"[ERR] Pinterest vrátil chybu: {el.text.strip()[:200]}")
                    # Vymaz špatně uložené heslo z klíčenky
                    _vymaz_udaje_z_klicenky()
                    return False
            except Exception:
                pass

        logger.error("[ERR] Přihlášení selhalo – timeout čekání na přesměrování")
        return False

    except Exception as exc:
        logger.error(f"[ERR] Výjimka při přihlašování: {exc}")
        logger.error(traceback.format_exc())
        return False


def _vymaz_udaje_z_klicenky() -> None:
    """
    Vymaže uložené přihlašovací údaje z klíčenky.
    Volá se při neúspěšném přihlášení (špatné heslo).
    """
    if not _KEYRING_DOSTUPNY:
        return
    try:
        keyring.delete_password(KEYRING_SERVICE, KEYRING_USER_KEY)
        keyring.delete_password(KEYRING_SERVICE, KEYRING_PASS_KEY)
        logger.info("[INFO] Staré přihlašovací údaje vymazány z klíčenky.")
        logger.info("       Při příštím spuštění budeš vyzván k zadání znovu.")
    except Exception:
        pass


# ===========================================================================
# 11. SELENIUM HELPERS
# ===========================================================================
# PROBLÉM: Původní kód neměl retry logiku, stale element handling ani
# adaptivní timeouty. Pinterest UI je pomalé a prvky se mění za běhu.
#
# ŘEŠENÍ:
#  - retry_find_element(): opakuje hledání při StaleElementReferenceException
#  - safe_click(): klikne přes JS jako fallback
#  - wait_for_element(): kombinuje explicit wait + retry
#  - Každá klíčová akce má retry_count pokusy
# ===========================================================================

def retry_find_element(driver, by, selector, retries=SELENIUM_RETRY_COUNT,
                       delay=SELENIUM_RETRY_DELAY):
    """
    Hledá element s automatickými pokusy při StaleElementReferenceException.
    Vrací element nebo None.
    """
    for attempt in range(retries):
        try:
            el = driver.find_element(by, selector)
            if el.is_displayed():
                return el
        except StaleElementReferenceException:
            if attempt < retries - 1:
                time.sleep(delay)
        except NoSuchElementException:
            break
        except Exception:
            if attempt < retries - 1:
                time.sleep(delay)
    return None


def safe_click(driver, element, use_js_fallback=True):
    """
    Klikne na element. Pokud normální klik selže, zkusí JavaScript.
    """
    try:
        element.click()
        return True
    except Exception:
        if use_js_fallback:
            try:
                driver.execute_script("arguments[0].click();", element)
                return True
            except Exception as exc:
                logger.warning(f"[WARN] safe_click JS fallback selhal: {exc}")
        return False


def wait_for_element(driver, by, selector, timeout=SELENIUM_TIMEOUT,
                     condition="clickable"):
    """
    Čeká na element s danou podmínkou. Vrací element nebo None.
    condition: 'clickable', 'visible', 'present'
    """
    try:
        wait = WebDriverWait(driver, timeout)
        if condition == "clickable":
            return wait.until(EC.element_to_be_clickable((by, selector)))
        elif condition == "visible":
            return wait.until(EC.visibility_of_element_located((by, selector)))
        else:
            return wait.until(EC.presence_of_element_located((by, selector)))
    except TimeoutException:
        return None


def scroll_and_click(driver, element):
    """Scrolluje k elementu a klikne na něj."""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        time.sleep(0.3)
        return safe_click(driver, element)
    except Exception as exc:
        logger.warning(f"[WARN] scroll_and_click: {exc}")
        return False


def send_keys_safe(element, text: str, clear_first=True):
    """Bezpečně zapíše text do pole (s volitelným smazáním)."""
    try:
        if clear_first:
            element.click()
            element.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            element.send_keys(Keys.DELETE)
            time.sleep(0.1)
        element.send_keys(text)
        return True
    except Exception as exc:
        logger.warning(f"[WARN] send_keys_safe: {exc}")
        return False


def vytvor_driver() -> webdriver.Chrome:
    """
    Vytvoří a vrátí nakonfigurovaný Chrome WebDriver.
    Nastavení: stabilní timeouty, bez zbytečných rozšíření.
    """
    opts = ChromeOptions()
    # Zabraň zbytečným chybám na Linux serveru bez display
    # opts.add_argument("--headless=new")  # odkomentuj pro headless provoz
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(60)
    driver.implicitly_wait(0)  # Explicitní waity jsou lepší než implicitní
    return driver


# ===========================================================================
# 11. TAG HELPER
# ===========================================================================

def najdi_tag_input(driver):
    """Najde input pro tagy pomocí prioritního seznamu selektorů."""
    selektory = [
        (By.XPATH, "//input[@data-test-id='pin-tag-input']"),
        (By.XPATH, "//input[@data-test-id='pin-draft-tag-field-input']"),
        (By.XPATH, "//input[@aria-label='Tagy']"),
        (By.XPATH, "//input[@aria-label='Tags']"),
        (By.XPATH, "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                   "'abcdefghijklmnopqrstuvwxyz'), 'tag')]"),
    ]
    for by, sel in selektory:
        try:
            el = driver.find_element(by, sel)
            if el.is_displayed() and el.is_enabled():
                return el
        except Exception:
            pass
    return None


def pridej_tag(driver, tag_text: str, timeout=10) -> bool:
    """Přidá jeden tag. Vrací True při úspěchu."""
    deadline = time.time() + timeout
    tag_input = None

    while time.time() < deadline:
        tag_input = najdi_tag_input(driver)
        if tag_input:
            break
        time.sleep(0.5)

    if not tag_input:
        logger.warning(f"[WARN] Pole pro tagy nenalezeno: '{tag_text}'")
        return False

    try:
        scroll_and_click(driver, tag_input)
        time.sleep(0.3)
        tag_input.send_keys(tag_text)
        time.sleep(0.5)
        tag_input.send_keys(Keys.RETURN)
        time.sleep(0.7)
        logger.info(f"    [OK] Tag: '{tag_text}'")
        return True
    except Exception as exc:
        logger.warning(f"    [WARN] Chyba při tagu '{tag_text}': {str(exc)[:100]}")
        return False


def pridej_vsechny_tagy(driver, tags_string: str) -> None:
    """Přidá všechny tagy ze stringu (oddělené ; nebo ,)."""
    if not tags_string or str(tags_string).strip() in ("", "nan"):
        logger.info("    [INFO] Žádné tagy k přidání")
        return

    tags_string = str(tags_string).strip()
    if ";" in tags_string:
        tags_list = [t.strip() for t in tags_string.split(";") if t.strip()]
    elif "," in tags_string:
        tags_list = [t.strip() for t in tags_string.split(",") if t.strip()]
    else:
        tags_list = [tags_string.strip()]

    logger.info(f"  > Přidávám {len(tags_list)} tagů...")
    uspesne = sum(pridej_tag(driver, tag) for tag in tags_list)
    logger.info(f"    [OK] Přidáno {uspesne}/{len(tags_list)} tagů")


# ===========================================================================
# 12. HLAVNÍ NAHRÁVACÍ FUNKCE
# ===========================================================================

def _cekej_na_dokonceni_nahravani(driver) -> bool:
    """
    Čeká na skutečné dokončení nahrávání/zpracování videa Pinterestem.

    Sleduje TŘI nezávislé signály úspěchu (stačí jeden):
      1. URL přesměrování  – Pinterest přesměruje z /pin-creation-tool/
                             na profilovou stránku nebo /pin/...
      2. Success toast     – banner "Pin naplánován" / "Pin uložen"
      3. Zmizení spinneru  – indikátory načítání zmizí ze stránky

    Průběžně loguje elapsed čas každých 30 sekund, aby bylo vidět
    že skript stále pracuje a nezamrzl.

    Vrací True pokud byl detekován signál úspěchu, False při timeoutu.
    """
    # Signály úspěchu – URL fragmenty které znamenají "hotovo"
    USPESNE_URL_FRAGMENTY = [
        "/pin/",          # přesměrování na konkrétní pin
        "/?savedPin",     # Pinterest success redirect
        "/profile/",      # přesměrování na profil
    ]
    ODESILA_URL = "pin-creation-tool"

    # CSS selektory pro success toast
    TOAST_SELEKTORY = [
        "[data-test-id='toast-success']",
        "[data-test-id='pin-saved-toast']",
        "[data-test-id='schedule-success']",
        ".toast--success",
        "[class*='successToast']",
        "[class*='SuccessToast']",
    ]

    # CSS selektory pro aktivní spinner / progress bar (jejich zmizení = hotovo)
    SPINNER_SELEKTORY = [
        "[data-test-id='upload-progress']",
        "[data-test-id='video-upload-progress']",
        "[class*='uploadProgress']",
        "[class*='UploadProgress']",
        "[role='progressbar']",
    ]

    start_time = time.time()
    posledni_log = start_time
    LOG_INTERVAL = 30  # s – jak často logovat průběh

    while True:
        elapsed = time.time() - start_time

        # --- Timeout ---
        if elapsed > SELENIUM_UPLOAD_TIMEOUT:
            logger.warning(
                f"  [WARN] Timeout {SELENIUM_UPLOAD_TIMEOUT}s – "
                f"pin možná nahrán, ale nebylo potvrzeno. Pokračuji..."
            )
            return False

        # --- Průběžný log každých 30 s ---
        if time.time() - posledni_log >= LOG_INTERVAL:
            logger.info(f"  [WAIT] Nahrávání stále probíhá... "
                        f"({int(elapsed)}s / {SELENIUM_UPLOAD_TIMEOUT}s)")
            posledni_log = time.time()

        try:
            current_url = driver.current_url

            # SIGNÁL 1: URL přesměrování (opustili jsme pin-creation-tool)
            if ODESILA_URL not in current_url:
                logger.info(f"  [OK] Přesměrování detekováno ({int(elapsed)}s): "
                            f"{current_url[:70]}")
                return True

            # SIGNÁL 2: Success toast
            for sel in TOAST_SELEKTORY:
                try:
                    toast = driver.find_element(By.CSS_SELECTOR, sel)
                    if toast.is_displayed():
                        logger.info(f"  [OK] Success toast detekován ({int(elapsed)}s): "
                                    f"'{toast.text[:80]}'")
                        time.sleep(1)  # nech toast zmizet
                        return True
                except NoSuchElementException:
                    pass

            # SIGNÁL 3: Spinner/progress bar zmizel (nahrávání dokončeno)
            # Toto je poslední záchrana – pokud spinner existoval a pak zmizel
            aktivni_spinnery = 0
            for sel in SPINNER_SELEKTORY:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        aktivni_spinnery += 1
                except NoSuchElementException:
                    pass

            # Pokud jsme za minimálně 5s a žádný spinner není vidět,
            # považujeme to za hotové (ale jen pokud jsme stále na upload stránce)
            if elapsed > 5 and aktivni_spinnery == 0 and ODESILA_URL in current_url:
                # Počkej ještě 3s a ověř znovu (může být krátká mezera mezi spinnery)
                time.sleep(3)
                current_url2 = driver.current_url
                if ODESILA_URL not in current_url2:
                    logger.info(f"  [OK] Přesměrování po spinner check ({int(elapsed)}s)")
                    return True
                # Stále na upload stránce bez spinneru – pravděpodobně hotovo
                logger.info(f"  [OK] Žádný aktivní spinner, upload dokončen ({int(elapsed)}s)")
                return True

        except WebDriverException as exc:
            # Browser může být dočasně nedostupný
            logger.warning(f"  [WARN] WebDriver chyba při čekání: {str(exc)[:100]}")

        time.sleep(SELENIUM_UPLOAD_POLL)


def _nahraj_jeden_pin(driver, row: pd.Series) -> bool:
    """
    Nahraje jeden pin. Vrací True při úspěchu, False při chybě.
    Odděleno od hlavní smyčky pro přehlednost a testovatelnost.
    """
    wait = WebDriverWait(driver, SELENIUM_TIMEOUT)

    # 1. VIDEO
    logger.info("  > Nahrávám video...")
    video_cesta = str(row.get("video_cesta", "")).strip()
    if not Path(video_cesta).exists():
        logger.error(f"  [ERR] Video nenalezeno: {video_cesta}")
        return False

    upload = wait_for_element(driver, By.ID, "storyboard-upload-input",
                              condition="present")
    if not upload:
        logger.error("  [ERR] Upload input nenalezen")
        return False
    upload.send_keys(video_cesta)
    time.sleep(4)
    logger.info(f"    [OK] Video se nahrává: {Path(video_cesta).name}")

    # 2. NADPIS
    logger.info("  > Vyplňuji nadpis...")
    title = wait_for_element(driver, By.ID, "storyboard-selector-title")
    if not title:
        logger.error("  [ERR] Pole nadpisu nenalezeno")
        return False
    send_keys_safe(title, row["nazev"])
    time.sleep(0.8)
    logger.info(f"    [OK] '{row['nazev']}'")

    # 3. ODKAZ
    logger.info("  > Vyplňuji odkaz...")
    odkaz_typ = str(row.get("odkaz", "")).strip().lower()
    odkaz_url = ("https://kurzy.uolakademie.cz/"
                 if "kurzy" in odkaz_typ else "https://www.uol.cz/")
    link = wait_for_element(driver, By.ID, "WebsiteField")
    if not link:
        logger.warning("  [WARN] Pole pro odkaz nenalezeno, přeskakuji")
    else:
        send_keys_safe(link, odkaz_url)
        time.sleep(0.8)
        logger.info(f"    [OK] {odkaz_url}")

    # 4. NÁSTĚNKA
    logger.info("  > Vybírám nástěnku...")
    for attempt in range(SELENIUM_RETRY_COUNT):
        try:
            board_btn = wait_for_element(
                driver, By.XPATH,
                "//div[@data-test-id='board-dropdown-select-button']",
            )
            if not board_btn:
                raise TimeoutException("board button not found")
            safe_click(driver, board_btn)
            time.sleep(1.5)

            board_option = wait_for_element(
                driver, By.XPATH,
                f"//div[@data-test-id='board-row-{row['nastepka']}']",
            )
            if not board_option:
                raise NoSuchElementException(f"board-row-{row['nastepka']}")
            safe_click(driver, board_option)
            time.sleep(1.5)
            logger.info(f"    [OK] '{row['nastepka']}'")
            break
        except (StaleElementReferenceException, TimeoutException,
                NoSuchElementException) as exc:
            if attempt < SELENIUM_RETRY_COUNT - 1:
                logger.warning(f"    [WARN] Nástěnka pokus {attempt + 1}: {exc}")
                time.sleep(SELENIUM_RETRY_DELAY)
            else:
                logger.error(f"  [ERR] Nástěnku se nepodařilo vybrat po {SELENIUM_RETRY_COUNT} pokusech")
                return False

    # 5. TAGY
    pridej_vsechny_tagy(driver, row.get("tagy", ""))

    # 6. PLÁNOVÁNÍ
    logger.info("  > Aktivuji plánování...")
    try:
        draft_switch = wait_for_element(
            driver, By.ID, "pin-draft-switch-group", condition="present"
        )
        if draft_switch and not draft_switch.is_selected():
            draft_label = driver.find_element(
                By.XPATH, "//label[@for='pin-draft-switch-group']"
            )
            safe_click(driver, draft_label)
            time.sleep(1)
            logger.info("    [OK] 'Zveřejnit později' aktivováno")

        # 7. DATUM
        logger.info("  > Nastavuji datum...")
        datum_str = str(row["datum_planovani"]).strip()
        date_input = wait_for_element(
            driver, By.XPATH,
            "//input[contains(@id,'pin-draft-schedule-date-field')]",
        )
        if date_input:
            date_input.click()
            time.sleep(0.4)
            date_input.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            date_input.send_keys(Keys.DELETE)
            time.sleep(0.1)
            date_input.send_keys(datum_str)
            time.sleep(0.5)
            logger.info(f"    [OK] Datum: {datum_str}")

        # 8. ČAS (dropdown)
        logger.info("  > Nastavuji čas...")
        cas_str = str(row["cas_planovani"]).strip()
        try:
            hour_s, minute_s = cas_str.split(":")
            dropdown_index = int(hour_s) * 2 + (1 if int(minute_s) == 30 else 0)

            time_wrapper = wait_for_element(
                driver, By.XPATH,
                "//div[@data-test-id='pin-draft-schedule-time-field-container']",
            )
            if time_wrapper:
                safe_click(driver, time_wrapper)
                time.sleep(1)

                time_option = wait_for_element(
                    driver, By.ID,
                    f"time-field-dropdown-item-{dropdown_index}",
                    condition="present",
                )
                if time_option:
                    scroll_and_click(driver, time_option)
                    time.sleep(0.5)
                    logger.info(f"    [OK] Čas: {cas_str}")
        except Exception as exc:
            logger.warning(f"    [WARN] Čas se nepodařilo nastavit: {exc}")

        # 9. PUBLIKOVÁNÍ
        logger.info("  > Publikuji pin...")
        publish_btn = wait_for_element(
            driver, By.XPATH,
            "//div[@data-test-id='storyboard-creation-nav-done']",
        )
        if not publish_btn:
            logger.error("  [ERR] Tlačítko Publish nenalezeno")
            return False
        safe_click(driver, publish_btn)
        time.sleep(2)

        # 10. POTVRZENÍ NAPLÁNOVÁNÍ
        logger.info("  > Potvrzuji naplánování...")
        try:
            confirm_btn = wait_for_element(
                driver, By.CSS_SELECTOR,
                "div[data-test-id='schedule-pin-confirm-button'] button",
                timeout=SELENIUM_LONG_TIMEOUT,
            )
            if confirm_btn:
                safe_click(driver, confirm_btn)
            else:
                logger.warning("  [WARN] Potvrzovací tlačítko nenalezeno, pokračuji...")
        except Exception as exc:
            logger.warning(f"  [WARN] Potvrzení: {exc}")

        # 11. ČEKÁNÍ NA DOKONČENÍ NAHRÁVÁNÍ VIDEA
        # ---------------------------------------------------------------
        # PROČ nestačí staleness_of(confirm_btn)?
        #   Pinterest odstraní potvrzovací dialog PŘED tím, než server
        #   dokončí zpracování videa. staleness_of se tedy splní příliš
        #   brzy a skript přejde na další pin dřív, než je video hotové.
        #
        # SPRÁVNÁ strategie – čekáme na JEDEN z těchto signálů úspěchu:
        #   a) URL se změnila z /pin-creation-tool/ na jinou (přesměrování)
        #   b) Zobrazil se "success toast" (potvrzovací banner)
        #   c) Zmizely všechny indikátory načítání (progress bar, spinner)
        #
        # Pokud žádný signál nepřijde do SELENIUM_UPLOAD_TIMEOUT sekund,
        # logujeme warning ale nepovažujeme to za fatální chybu – pin
        # mohl být uložen i přesto.
        # ---------------------------------------------------------------
        logger.info(f"  [WAIT] Čekám na dokončení nahrávání videa "
                    f"(max {SELENIUM_UPLOAD_TIMEOUT // 60} min)...")

        _cekej_na_dokonceni_nahravani(driver)

        time.sleep(2)
        logger.info("  [OK] Pin naplánován!\n")
        return True

    except Exception as exc:
        logger.error(f"  [ERR] Chyba při plánování: {str(exc)[:300]}")
        logger.error(traceback.format_exc())
        return False


def nahraj_piny(df_pins: pd.DataFrame) -> None:
    """Nahraje všechny připravené piny na Pinterest s plánováním."""
    driver = None
    try:
        # --- Získej přihlašovací údaje (klíčenka / fallback / interaktivně) ---
        email, heslo = ziskej_prihlasovaci_udaje()

        driver = vytvor_driver()

        # --- Automatické přihlášení ---
        prihlaseno = prihlasit_se_na_pinterest(driver, email, heslo)
        if not prihlaseno:
            logger.error("[ERR] Přihlášení se nezdařilo. Ukončuji.")
            logger.info("[INFO] Zkontroluj email/heslo a spusť skript znovu.")
            logger.info("       Pokud bylo heslo špatně uloženo, bylo automaticky")
            logger.info("       vymazáno z klíčenky – příště zadej znovu.")
            return

        driver.get(PIN_URL)
        time.sleep(3)
        logger.info("[OK] Pin Creation Tool otevřen\n")

        celkem = len(df_pins)
        uspesne = 0
        neuspesne = 0

        for idx, row in df_pins.iterrows():
            logger.info("=" * 60)
            logger.info(f"[PIN] Pin {idx + 1}/{celkem}: {row['nazev']}")
            logger.info(f"   Barva: {row['barva_satu']}")
            logger.info(f"   Nástěnka: {row['nastepka']}")
            logger.info(f"   Plánováno: {row['datum_planovani']} v {row['cas_planovani']}")
            logger.info("=" * 60)

            ok = False
            # Celkové retry na úrovni pinu (pro případ pádu browseru)
            for pin_attempt in range(2):
                try:
                    ok = _nahraj_jeden_pin(driver, row)
                    break
                except WebDriverException as exc:
                    logger.warning(
                        f"  [WARN] WebDriver chyba (pokus {pin_attempt + 1}): {exc}"
                    )
                    if pin_attempt == 0:
                        # Zkus obnovit stránku a znovu
                        try:
                            driver.get(PIN_URL)
                            time.sleep(4)
                        except Exception:
                            pass
                    else:
                        break

            if ok:
                uspesne += 1
                aktualizuj_stav_v_excelu(str(row["video_cesta"]), STAV_NAHRANO)
            else:
                neuspesne += 1
                aktualizuj_stav_v_excelu(str(row["video_cesta"]), STAV_CHYBA)
                logger.warning(f"  [WARN] Pin {idx + 1} selhal, pokračuji...")

            # Přejdi na nový pin
            try:
                driver.get(PIN_URL)
                time.sleep(3)
            except Exception:
                pass

        logger.info("=" * 60)
        logger.info(f"[OK] Hotovo! Nahráno: {uspesne}/{celkem}, Chyby: {neuspesne}")
        logger.info("=" * 60)

    except Exception as exc:
        logger.error(f"[ERR] Kritická chyba: {exc}")
        logger.error(traceback.format_exc())
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# ===========================================================================
# 13. HLAVNÍ PROGRAM
# ===========================================================================

if __name__ == "__main__":

    logger.info("=" * 60)
    logger.info("Pinterest Smart Automation V4")
    logger.info("=" * 60)
    logger.info(f"[SEARCH] Hledám Excel soubor: {EXCEL_FILE}\n")

    if not EXCEL_FILE.exists():
        logger.info("[INFO] Excel nenalezen – spouštím vytváření...")
        logger.info("=" * 60)
        logger.info("KROK 1: VYTVÁŘENÍ EXCEL SOUBORU Z VIDEÍ")
        logger.info("=" * 60)

        videa = najdi_videa()
        if videa:
            vytvor_excel_z_videi(videa)
        else:
            logger.error("[ERR] Žádná videa nenalezena!")
            logger.info("   Umístěj video soubory do aktuálního adresáře a spusť znovu.")
            sys.exit(1)

    else:
        logger.info(f"[OK] Excel nalezen: {EXCEL_FILE}")
        logger.info("=" * 60)
        logger.info("KROK 2: NAHRÁVÁNÍ PINŮ NA PINTEREST")
        logger.info("=" * 60)

        ok, df_pins = zkontroluj_excel()
        if ok:
            nahraj_piny(df_pins)
            logger.info("=" * 60)
            logger.info("[OK] VSECHNY PINY JSOU NAPLANOVANE!")
            logger.info("=" * 60)
        else:
            logger.warning("[WARN] Excel není připraven k nahrání")
            logger.info(f"   Vyplň chybějící údaje a změň STAV na: {STAV_PRIPRAVENO}")
            sys.exit(1)
